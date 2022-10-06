from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Ram": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Shyam": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Raj": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"AWS": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Navya": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Ram'].keys())
ws.append(headings)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['Ram']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("UpdatedGrades.xlsx")